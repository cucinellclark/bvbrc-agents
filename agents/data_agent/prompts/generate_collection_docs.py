#!/usr/bin/env python3
"""
Generate the collection documentation section of the system prompt from
data_types.xlsx.

Reads every sheet in the spreadsheet and produces a compact, LLM-friendly
reference that is written to `collection_reference.py` as a Python string
constant.

Usage (from the Data/ directory):
    python -m data_agent.prompts.generate_collection_docs

The output file (collection_reference.py) is committed to the repo so the
agent can run without openpyxl or the xlsx at runtime.
"""

from __future__ import annotations

import os
import textwrap
from pathlib import Path

# ---------------------------------------------------------------------------
# Fields to exclude from the prompt (internal / system fields)
# ---------------------------------------------------------------------------
EXCLUDED_FIELDS = {
    "version",
    "owner",
    "public",
    "user_read",
    "user_write",
    "date_inserted",
    "date_modified",
    "_version_",
}

# Where to find the spreadsheet (relative to repo root)
SPREADSHEET_NAME = "data_types.xlsx"


def _find_spreadsheet() -> Path:
    """Locate data_types.xlsx relative to this script."""
    # Script is at data_agent/prompts/generate_collection_docs.py
    # Spreadsheet is at Data/data_types.xlsx (two levels up)
    script_dir = Path(__file__).resolve().parent
    candidates = [
        script_dir.parent.parent / SPREADSHEET_NAME,  # Data/data_types.xlsx
        script_dir.parent / SPREADSHEET_NAME,          # data_agent/data_types.xlsx
        Path.cwd() / SPREADSHEET_NAME,                 # current dir
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(
        f"Cannot find {SPREADSHEET_NAME}. Searched: {[str(c) for c in candidates]}"
    )


def _parse_spreadsheet(path: Path) -> tuple[list[dict], dict[str, list[dict]]]:
    """
    Parse the spreadsheet and return:
      - index: list of {name, purpose} dicts from the data_types sheet
      - collections: dict mapping collection name -> list of field dicts
        Each field dict has: name, type, definition
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to generate collection docs. "
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(str(path), read_only=True)

    # Parse the index sheet
    index = []
    ws_index = wb["data_types"]
    seen_names = set()
    for row in ws_index.iter_rows(min_row=2, values_only=True):
        name = row[0]
        purpose = row[1]
        if name and name not in seen_names:
            seen_names.add(name)
            index.append({"name": str(name).strip(), "purpose": str(purpose).strip()})

    # Parse each collection sheet
    collections: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "data_types":
            continue

        ws = wb[sheet_name]
        metadata: dict[str, str] = {}
        fields: list[dict] = []
        in_fields = False

        for row in ws.iter_rows(values_only=True):
            vals = list(row)

            # Parse metadata rows (before the field table)
            if vals[0] in ("Data type", "Primary key", "Purpose"):
                metadata[str(vals[0])] = str(vals[1]).strip() if vals[1] else ""
                continue

            # Detect the header row of the field table
            if vals[0] == "Attribute name":
                in_fields = True
                continue

            # Parse field rows
            if in_fields and vals[0] is not None:
                field_name = str(vals[0]).strip().rstrip("*")  # remove PK marker
                field_type = str(vals[1]).strip() if vals[1] else "string"
                definition = str(vals[2]).strip() if vals[2] else ""

                if field_name.lower() in EXCLUDED_FIELDS:
                    continue

                fields.append({
                    "name": field_name,
                    "type": field_type,
                    "definition": definition,
                })

        pk = metadata.get("Primary key", "")
        collections[sheet_name] = {
            "primary_key": pk,
            "fields": fields,
        }

    wb.close()
    return index, collections


def _format_collection_reference(
    index: list[dict],
    collections: dict[str, dict],
) -> str:
    """Format the parsed data into a compact LLM-friendly reference string."""
    lines = []

    # Section 1: Collection index
    lines.append("=== AVAILABLE COLLECTIONS ===")
    lines.append("")
    for entry in index:
        lines.append(f"- {entry['name']}: {entry['purpose']}")
    lines.append("")

    # Section 2: Per-collection field reference
    lines.append("=== COLLECTION FIELD REFERENCE ===")
    lines.append("")
    lines.append(
        "Below are the queryable fields for each collection. Use ONLY these field "
        "names in your Solr queries. If a field is not listed for a collection, "
        "it does NOT exist there -- do not guess or assume fields exist across "
        "collections."
    )
    lines.append("")

    for entry in index:
        cname = entry["name"]
        cdata = collections.get(cname)
        if not cdata:
            continue

        pk = cdata["primary_key"]
        fields = cdata["fields"]

        lines.append(f"[{cname}]  (primary key: {pk})")
        for f in fields:
            # Compact format: field_name (type) -- definition
            defn = f["definition"]
            # Truncate very long definitions
            if len(defn) > 120:
                defn = defn[:117] + "..."
            lines.append(f"  {f['name']} ({f['type']}): {defn}")
        lines.append("")

    return "\n".join(lines)


def generate() -> str:
    """Main entry point: parse spreadsheet and return formatted reference."""
    xlsx_path = _find_spreadsheet()
    print(f"Reading: {xlsx_path}")
    index, collections = _parse_spreadsheet(xlsx_path)
    reference = _format_collection_reference(index, collections)

    total_fields = sum(len(c["fields"]) for c in collections.values())
    print(f"Parsed {len(index)} collections, {total_fields} fields (excluding system fields)")
    print(f"Reference text: {len(reference)} chars, ~{len(reference) // 4} tokens (approx)")

    return reference


def write_module(reference: str) -> Path:
    """Write the reference string as a Python module."""
    output_path = Path(__file__).resolve().parent / "collection_reference.py"

    # Escape the reference for embedding in a triple-quoted string
    # Only need to escape backslashes and triple-quotes
    escaped = reference.replace("\\", "\\\\").replace('"""', '\\"\\"\\"')

    content = (
        '"""Auto-generated collection reference for the system prompt.\n'
        "\n"
        "DO NOT EDIT THIS FILE MANUALLY.\n"
        "Regenerate with: python -m data_agent.prompts.generate_collection_docs\n"
        '"""\n'
        "\n"
        f'COLLECTION_REFERENCE = """\\\n{escaped}"""\n'
    )

    output_path.write_text(content)
    print(f"Wrote: {output_path}")
    return output_path


if __name__ == "__main__":
    ref = generate()
    write_module(ref)
    print("Done.")
