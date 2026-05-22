"""
Tool implementations for collection introspection.

These tools use data parsed from data_types.xlsx rather than the MCP server's
prompt-file-based implementations (data_functions.list_solr_collections and
data_functions.get_collection_fields). The spreadsheet provides richer metadata
(field types, definitions, examples) that the MCP prompt files do not have.

The MCP server equivalents are:
  - data_functions.list_solr_collections() -> static text string
  - data_functions.get_collection_fields() -> list of field names only

Our versions return structured dicts with types and definitions, which is
more useful for both the LLM and programmatic consumers.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Dict, List, Optional

# ---------------------------------------------------------------------------
# Load collection metadata from the spreadsheet (cached at module level)
# ---------------------------------------------------------------------------

_COLLECTION_DATA: Optional[Dict[str, Dict[str, Any]]] = None
_COLLECTION_INDEX: Optional[List[Dict[str, str]]] = None


def _load_collection_data() -> tuple[list[dict[str, str]], dict[str, dict[str, Any]]]:
    """
    Parse data_types.xlsx and cache the results. Returns (index, collections).

    The index is a list of {name, purpose} dicts.
    The collections dict maps collection name -> {primary_key, fields: [{name, type, definition}]}.
    """
    global _COLLECTION_DATA, _COLLECTION_INDEX

    if _COLLECTION_DATA is not None and _COLLECTION_INDEX is not None:
        return _COLLECTION_INDEX, _COLLECTION_DATA

    try:
        import openpyxl
    except ImportError:
        # If openpyxl is not available, return empty data
        _COLLECTION_INDEX = []
        _COLLECTION_DATA = {}
        return _COLLECTION_INDEX, _COLLECTION_DATA

    # Find the spreadsheet
    xlsx_path = _find_spreadsheet()
    if xlsx_path is None:
        _COLLECTION_INDEX = []
        _COLLECTION_DATA = {}
        return _COLLECTION_INDEX, _COLLECTION_DATA

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)

    # Parse index sheet
    index = []
    ws_index = wb["data_types"]
    seen_names = set()
    for row in ws_index.iter_rows(min_row=2, values_only=True):
        name = row[0]
        purpose = row[1]
        if name and name not in seen_names:
            seen_names.add(name)
            index.append({"name": str(name).strip(), "purpose": str(purpose).strip()})

    # Excluded system fields
    excluded = {
        "version", "owner", "public", "user_read", "user_write",
        "date_inserted", "date_modified", "_version_",
    }

    # Parse each collection sheet
    collections: Dict[str, Dict[str, Any]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "data_types":
            continue

        ws = wb[sheet_name]
        metadata: Dict[str, str] = {}
        fields: List[Dict[str, str]] = []
        in_fields = False

        for row in ws.iter_rows(values_only=True):
            vals = list(row)

            if vals[0] in ("Data type", "Primary key", "Purpose"):
                metadata[str(vals[0])] = str(vals[1]).strip() if vals[1] else ""
                continue

            if vals[0] == "Attribute name":
                in_fields = True
                continue

            if in_fields and vals[0] is not None:
                field_name = str(vals[0]).strip().rstrip("*")
                if field_name.lower() in excluded:
                    continue

                field_type = str(vals[1]).strip() if vals[1] else "string"
                definition = str(vals[2]).strip() if vals[2] else ""

                fields.append({
                    "name": field_name,
                    "type": field_type,
                    "definition": definition,
                })

        collections[sheet_name] = {
            "primary_key": metadata.get("Primary key", ""),
            "purpose": metadata.get("Purpose", ""),
            "fields": fields,
        }

    wb.close()

    _COLLECTION_INDEX = index
    _COLLECTION_DATA = collections
    return _COLLECTION_INDEX, _COLLECTION_DATA


def _find_spreadsheet() -> Optional[Path]:
    """Locate data_types.xlsx."""
    candidates = [
        Path(__file__).resolve().parent.parent.parent / "data_types.xlsx",
        Path.cwd() / "data_types.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


# ---------------------------------------------------------------------------
# Tool functions
# ---------------------------------------------------------------------------

async def list_collections() -> Dict[str, Any]:
    """
    List all available BV-BRC Solr collections with descriptions.

    Returns:
        Dict with "collections" key containing a list of {name, purpose} dicts.
    """
    index, _ = _load_collection_data()

    if not index:
        # Fallback: return the static list from the tool registry
        from data_agent.tool_registry import COLLECTIONS
        return {
            "collections": [{"name": c, "purpose": ""} for c in COLLECTIONS],
            "count": len(COLLECTIONS),
        }

    return {
        "collections": index,
        "count": len(index),
    }


async def get_collection_fields(collection: str) -> Dict[str, Any]:
    """
    Get the queryable fields for a specific BV-BRC Solr collection.

    Args:
        collection: The collection name to inspect.

    Returns:
        Dict with "fields" key containing a list of {name, type, definition} dicts,
        or an error if the collection is not found.
    """
    _, collections = _load_collection_data()

    cdata = collections.get(collection)
    if cdata is None:
        return {
            "error": f"Collection '{collection}' not found.",
            "available_collections": sorted(collections.keys()),
        }

    return {
        "collection": collection,
        "primary_key": cdata["primary_key"],
        "field_count": len(cdata["fields"]),
        "fields": cdata["fields"],
    }
